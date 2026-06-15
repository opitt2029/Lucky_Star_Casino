import sys
import openpyxl

PATH = 'docs/幸運星幣城_工作分配表.xlsx'
DONE = '✅ 已完成'
TARGETS = {'T-002', 'T-041', 'T-042', 'T-050', 'T-051', 'T-052', 'T-053', 'T-105', 'T-106'}

apply = '--apply' in sys.argv
wb = openpyxl.load_workbook(PATH)
changes = 0

for ws in wb.worksheets:
    # 找出含「任務編號」與「狀態」的標題列，定出兩欄索引（同一 sheet 欄位位置固定）
    id_col = status_col = None
    for r in range(1, ws.max_row + 1):
        row_vals = {ws.cell(row=r, column=c).value: c for c in range(1, ws.max_column + 1)}
        if '任務編號' in row_vals and '狀態' in row_vals:
            id_col = row_vals['任務編號']
            status_col = row_vals['狀態']
            break
    if id_col is None or status_col is None:
        print(f'[skip] {ws.title}: 無任務編號/狀態欄')
        continue

    for r in range(1, ws.max_row + 1):
        tid = ws.cell(row=r, column=id_col).value
        if isinstance(tid, str) and tid.strip() in TARGETS:
            cell = ws.cell(row=r, column=status_col)
            old = cell.value
            if old != DONE:
                print(f'[{ws.title}] row {r} {tid.strip()}: {old!r} -> {DONE!r}')
                changes += 1
                if apply:
                    cell.value = DONE

print(f'\nTOTAL changes: {changes} (apply={apply})')
if apply:
    wb.save(PATH)
    print('SAVED', PATH)
